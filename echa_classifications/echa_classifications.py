#!/usr/bin/env python3
"""Extract harmonised and industry (self-/notified) classifications for CAS numbers.

The two data types live in different ECHA sources, so this tool joins them locally
against a list of CAS numbers instead of scraping the (bot-protected) ECHA CHEM portal:

  * Harmonised classification  -> ECHA's downloadable "Annex VI, Table 3 to CLP" Excel
                                  https://echa.europa.eu/information-on-chemicals/annex-vi-to-clp
  * Industry classification    -> an export from the ECHA CHEM C&L Inventory
                                  https://chem.echa.europa.eu/  (search, then "Export")

Both inputs are files you download; this tool auto-detects their columns (ECHA changes
headers between releases), matches each CAS number, and writes a merged CSV. It prints the
column mapping it inferred so you can catch and override a mis-detection.

Run `python echa_classifications.py --help` for usage.
"""

import argparse
import csv
import io
import re
import sys
from collections import OrderedDict


# --------------------------------------------------------------------------------------
# CAS number handling
# --------------------------------------------------------------------------------------

# A CAS Registry Number is 2-7 digits, a 2-digit block, then a single check digit.
CAS_RE = re.compile(r"(\d{2,7})-(\d{2})-(\d)")


def canonical_cas(value):
    """Return the canonical "NNNN-NN-N" form of the first CAS number found, or "".

    Leading zeros are stripped from the first block so that "0000071-43-2" and
    "71-43-2" collapse to the same key. Returns "" if no CAS pattern is present.
    """
    if value is None:
        return ""
    m = CAS_RE.search(str(value))
    if not m:
        return ""
    return "{}-{}-{}".format(int(m.group(1)), m.group(2), m.group(3))


def all_cas(value):
    """Return every distinct canonical CAS number in a cell (some cells list several)."""
    if value is None:
        return []
    seen = []
    for m in CAS_RE.finditer(str(value)):
        c = "{}-{}-{}".format(int(m.group(1)), m.group(2), m.group(3))
        if c not in seen:
            seen.append(c)
    return seen


def cas_checksum_ok(cas):
    """Validate a canonical CAS via its check digit. None if it isn't CAS-shaped."""
    m = re.fullmatch(r"(\d+)-(\d{2})-(\d)", cas or "")
    if not m:
        return None
    body = m.group(1) + m.group(2)
    check = int(m.group(3))
    total = sum(int(d) * i for i, d in enumerate(reversed(body), start=1))
    return total % 10 == check


# --------------------------------------------------------------------------------------
# Reading tabular files (.xlsx via openpyxl, or delimited text)
# --------------------------------------------------------------------------------------

def read_rows(path):
    """Read a spreadsheet/CSV into a list of rows, each a list of stringified cells."""
    lower = path.lower()
    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        return _read_xlsx(path)
    if lower.endswith(".xls"):
        raise SystemExit(
            "ERROR: '{}' is an old-format .xls file, which cannot be read directly.\n"
            "       Open it and 'Save As' .xlsx or .csv, then pass that instead.".format(path)
        )
    return _read_delimited(path)


def _read_xlsx(path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit(
            "ERROR: reading '{}' needs the 'openpyxl' library.\n"
            "       Install it with:  pip install openpyxl\n"
            "       (or open the file and 'Save As' CSV, then pass the .csv).".format(path)
        )
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for raw in ws.iter_rows(values_only=True):
        rows.append(["" if c is None else str(c).strip() for c in raw])
    wb.close()
    return rows


def _read_delimited(path):
    # ECHA/European CSVs are sometimes semicolon-separated and Latin-1 encoded.
    data = None
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with open(path, "r", encoding=enc, newline="") as fh:
                data = fh.read()
            break
        except UnicodeDecodeError:
            continue
    if data is None:
        raise SystemExit("ERROR: could not decode '{}' as text.".format(path))

    sample = data[:8192]
    delimiter = ","
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        # Fall back to whichever common delimiter appears most in the header line.
        header = sample.splitlines()[0] if sample.splitlines() else ""
        delimiter = max(",;\t|", key=header.count)

    reader = csv.reader(io.StringIO(data), delimiter=delimiter)
    return [[("" if c is None else c.strip()) for c in row] for row in reader]


# --------------------------------------------------------------------------------------
# Header detection and column mapping
# --------------------------------------------------------------------------------------

def norm(text):
    """Normalise a header for fuzzy matching: lowercase, alphanumerics only."""
    return re.sub(r"[^a-z0-9]", "", str(text).lower())


def detect_header_row(rows, max_scan=40):
    """Find the header row: the first row (within max_scan) that carries a CAS column.

    ECHA sheets often have title/blurb rows above the real header, so we look for the
    row that actually contains a "CAS" column. Falls back to the first non-empty row.
    """
    for i, row in enumerate(rows[:max_scan]):
        for cell in row:
            n = norm(cell)
            if n in ("casno", "casnumber", "casrn", "cas") or (
                "cas" in n and ("no" in n or "number" in n or "rn" in n)
            ):
                return i
    for i, row in enumerate(rows):
        if any(str(c).strip() for c in row):
            return i
    return 0


def find_column(headers, aliases):
    """Return the index of the first header whose normalised form contains any alias."""
    nheaders = [norm(h) for h in headers]
    for i, nh in enumerate(nheaders):
        if any(a in nh for a in aliases):
            return i
    return None


# Alias groups (already normalised) used to locate columns in either source file.
CAS_ALIASES = ["casno", "casnumber", "casrn", "cas"]
EC_ALIASES = ["ecno", "ecnumber", "eclistno", "einecs", "elincs", "listno"]
NAME_ALIASES = [
    "internationalchemicalidentification", "chemicalname", "substancename",
    "substanceidentity", "iupacname", "name", "identification",
]
INDEX_ALIASES = ["indexno", "indexnumber", "index"]
HAZ_CLASS_ALIASES = ["hazardclassandcategory", "hazardclasscategory", "hazardclass"]
HAZ_STATEMENT_ALIASES = ["hazardstatementcode", "hazardstatement"]
PICTOGRAM_ALIASES = ["pictogramsignalword", "pictogram", "signalword"]
CONCLIMIT_ALIASES = ["specificconclimit", "concentrationlimit", "mfactor", "specconc"]
NOTES_ALIASES = ["notes", "note"]

# Any header hinting at classification content, used to pick up the industry columns
# generically (their exact names vary by ECHA release).
CLASSIFICATION_HINTS = [
    "hazard", "classification", "signalword", "pictogram", "notifier", "notified",
    "labelling", "labeling", "ghs", "precautionary", "hstatement", "pstatement",
    "concentrationlimit", "mfactor", "specificconc", "reason", "clp",
]


# --------------------------------------------------------------------------------------
# Source-specific indexers
# --------------------------------------------------------------------------------------

class HarmonisedIndex:
    """CAS -> harmonised (Annex VI Table 3) record, plus the detected column mapping."""

    FIELDS = [
        ("index_no", "harmonised_index_no"),
        ("name", "harmonised_name"),
        ("ec_number", "harmonised_ec_number"),
        ("hazard_class_category", "harmonised_hazard_class_and_category"),
        ("hazard_statements", "harmonised_hazard_statements"),
        ("pictogram_signal", "harmonised_pictogram_signal_word"),
        ("conc_limits", "harmonised_spec_conc_limits_m_factors"),
        ("notes", "harmonised_notes"),
    ]

    def __init__(self, path, cas_col_override=None):
        rows = read_rows(path)
        if not rows:
            raise SystemExit("ERROR: harmonised file '{}' is empty.".format(path))
        hidx = detect_header_row(rows)
        headers = rows[hidx]
        self.headers = headers
        data_rows = rows[hidx + 1:]

        cas_i = cas_col_override if cas_col_override is not None else find_column(headers, CAS_ALIASES)
        if cas_i is None:
            raise SystemExit(
                "ERROR: no CAS column found in harmonised file '{}'.\n"
                "       Detected headers: {}\n"
                "       Override with --harmonised-cas-col <0-based index>.".format(path, headers)
            )

        # Locate the remaining columns. The classification "Hazard statement code(s)"
        # column appears before the pictogram column; the labelling one after it, so we
        # take the first hazard-statement match, which is the classification column.
        col = {
            "cas": cas_i,
            "index_no": find_column(headers, INDEX_ALIASES),
            "name": find_column(headers, NAME_ALIASES),
            "ec_number": find_column(headers, EC_ALIASES),
            "hazard_class_category": find_column(headers, HAZ_CLASS_ALIASES),
            "hazard_statements": find_column(headers, HAZ_STATEMENT_ALIASES),
            "pictogram_signal": find_column(headers, PICTOGRAM_ALIASES),
            "conc_limits": find_column(headers, CONCLIMIT_ALIASES),
            "notes": find_column(headers, NOTES_ALIASES),
        }
        self.col = col
        self.mapping = OrderedDict(
            (k, headers[v] if v is not None and v < len(headers) else None)
            for k, v in col.items()
        )

        self.by_cas = {}
        self.rows_read = 0
        for row in data_rows:
            cas_cell = row[cas_i] if cas_i < len(row) else ""
            cas_list = all_cas(cas_cell)
            if not cas_list:
                continue
            self.rows_read += 1
            record = {}
            for key in col:
                if key == "cas":
                    continue
                ci = col[key]
                record[key] = row[ci].strip() if ci is not None and ci < len(row) else ""
            for cas in cas_list:
                # First entry wins; harmonised entries are effectively unique per CAS.
                self.by_cas.setdefault(cas, record)

    def get(self, cas):
        return self.by_cas.get(cas)


class IndustryIndex:
    """CAS -> industry (ECHA CHEM C&L export) record, with generic classification columns."""

    def __init__(self, path, cas_col_override=None, col_overrides=None):
        rows = read_rows(path)
        if not rows:
            raise SystemExit("ERROR: industry file '{}' is empty.".format(path))
        hidx = detect_header_row(rows)
        headers = rows[hidx]
        self.headers = headers
        data_rows = rows[hidx + 1:]

        cas_i = cas_col_override if cas_col_override is not None else find_column(headers, CAS_ALIASES)
        if cas_i is None:
            raise SystemExit(
                "ERROR: no CAS column found in industry export '{}'.\n"
                "       Detected headers: {}\n"
                "       Override with --industry-cas-col <0-based index>.".format(path, headers)
            )
        self.cas_i = cas_i

        name_i = find_column(headers, NAME_ALIASES)
        ec_i = find_column(headers, EC_ALIASES)

        # Which columns carry the actual classification content?
        if col_overrides:
            wanted = set(col_overrides)
            class_cols = [i for i, h in enumerate(headers) if h in wanted or str(i) in wanted]
        else:
            skip = {cas_i, name_i, ec_i}
            class_cols = [
                i for i, h in enumerate(headers)
                if i not in skip and any(hint in norm(h) for hint in CLASSIFICATION_HINTS)
            ]
        self.class_cols = class_cols
        # Output column labels for the industry side (stable across all rows).
        self.out_columns = ["industry_" + (headers[i] or "col{}".format(i)) for i in class_cols]
        if name_i is not None:
            self.out_columns.insert(0, "industry_name")
        self.name_i = name_i
        self.ec_i = ec_i

        self.mapping = OrderedDict()
        self.mapping["cas"] = headers[cas_i] if cas_i < len(headers) else None
        self.mapping["name"] = headers[name_i] if name_i is not None else None
        self.mapping["ec_number"] = headers[ec_i] if ec_i is not None else None
        self.mapping["classification_columns"] = [headers[i] for i in class_cols]

        self.by_cas = {}
        self.rows_read = 0
        self.duplicate_cas = set()
        for row in data_rows:
            cas_cell = row[cas_i] if cas_i < len(row) else ""
            cas_list = all_cas(cas_cell)
            if not cas_list:
                continue
            self.rows_read += 1
            record = OrderedDict()
            if name_i is not None:
                record["industry_name"] = row[name_i].strip() if name_i < len(row) else ""
            for i in class_cols:
                label = "industry_" + (headers[i] or "col{}".format(i))
                record[label] = row[i].strip() if i < len(row) else ""
            for cas in cas_list:
                if cas in self.by_cas:
                    self.duplicate_cas.add(cas)
                    continue  # keep first row for a CAS
                self.by_cas[cas] = record

    def get(self, cas):
        return self.by_cas.get(cas)


# --------------------------------------------------------------------------------------
# CAS list input
# --------------------------------------------------------------------------------------

def load_cas_list(path):
    """Load CAS numbers from a .txt (one per line) or .csv (auto-detect CAS column).

    Preserves the original strings so the output echoes exactly what the user supplied.
    """
    lower = path.lower()
    originals = []
    if lower.endswith((".csv", ".tsv", ".xlsx", ".xlsm")):
        rows = read_rows(path)
        if not rows:
            return originals
        hidx = detect_header_row(rows)
        headers = rows[hidx]
        cas_i = find_column(headers, CAS_ALIASES)
        if cas_i is None:
            cas_i = 0  # assume first column
            start = hidx  # no recognisable header; treat every row as data
        else:
            start = hidx + 1
        for row in rows[start:]:
            if cas_i < len(row) and str(row[cas_i]).strip():
                originals.append(str(row[cas_i]).strip())
    else:
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                with open(path, "r", encoding=enc) as fh:
                    for line in fh:
                        line = line.strip()
                        if not line or line.startswith("#"):
                            continue
                        originals.append(line)
                break
            except UnicodeDecodeError:
                originals = []
                continue
    return originals


# --------------------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------------------

def build_output_rows(cas_inputs, harmonised, industry):
    """Join the requested CAS numbers against both indexes into ordered dict rows."""
    industry_cols = industry.out_columns if industry else []
    out_rows = []
    stats = {"total": 0, "harmonised": 0, "industry": 0, "invalid": 0, "no_cas": 0}

    seen = set()
    for original in cas_inputs:
        cas = canonical_cas(original)
        stats["total"] += 1
        row = OrderedDict()
        row["cas_input"] = original
        row["cas_number"] = cas

        if not cas:
            stats["no_cas"] += 1
            row["cas_valid"] = "no-cas-found"
        else:
            ok = cas_checksum_ok(cas)
            row["cas_valid"] = "yes" if ok else ("no" if ok is False else "unknown")
            if ok is False:
                stats["invalid"] += 1

        # Always present so every row has the same columns for the CSV writer.
        note = ""
        if cas and cas in seen:
            note = "duplicate in input"
        elif cas:
            seen.add(cas)
        row["note"] = note

        # Only emit a source's columns when that source was actually provided.
        if harmonised is not None:
            h = harmonised.get(cas) if cas else None
            row["harmonised_found"] = "yes" if h else "no"
            for key, out_name in HarmonisedIndex.FIELDS:
                row[out_name] = h.get(key, "") if h else ""
            if h:
                stats["harmonised"] += 1

        if industry is not None:
            ind = industry.get(cas) if cas else None
            row["industry_found"] = "yes" if ind else "no"
            for col in industry_cols:
                row[col] = ind.get(col, "") if ind else ""
            if ind:
                stats["industry"] += 1

        out_rows.append(row)

    return out_rows, stats


def write_csv(path, rows):
    if not rows:
        # Still write a header-less empty file so the run has a defined output.
        open(path, "w").close()
        return
    fieldnames = list(rows[0].keys())
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for r in rows:
            writer.writerow(r)


def print_mapping(title, mapping):
    print("\n[{}] detected columns:".format(title), file=sys.stderr)
    for key, val in mapping.items():
        shown = val if val not in (None, [], "") else "(not found)"
        print("    {:<24} -> {}".format(key, shown), file=sys.stderr)


def parse_args(argv):
    p = argparse.ArgumentParser(
        description="Merge ECHA harmonised (Annex VI Table 3) and industry "
                    "(ECHA CHEM C&L export) classifications for a list of CAS numbers.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    src = p.add_argument_group("inputs")
    src.add_argument("--cas", nargs="+", metavar="CAS", help="CAS numbers to look up.")
    src.add_argument("--cas-list", metavar="FILE",
                     help="File of CAS numbers: .txt (one per line) or .csv/.xlsx (CAS column auto-detected).")
    src.add_argument("--harmonised", metavar="FILE",
                     help="ECHA 'Annex VI, Table 3 to CLP' Excel/CSV (harmonised classifications).")
    src.add_argument("--industry", metavar="FILE",
                     help="ECHA CHEM C&L Inventory export (industry self-/notified classifications).")
    out = p.add_argument_group("output")
    out.add_argument("--output", "-o", default="echa_classifications.csv", metavar="FILE",
                     help="Output CSV path (default: echa_classifications.csv).")
    ov = p.add_argument_group("column overrides (use if auto-detection maps a column wrong)")
    ov.add_argument("--harmonised-cas-col", type=int, metavar="N",
                    help="0-based index of the CAS column in the harmonised file.")
    ov.add_argument("--industry-cas-col", type=int, metavar="N",
                    help="0-based index of the CAS column in the industry export.")
    ov.add_argument("--industry-cols", metavar="COLS",
                    help="Comma-separated header names (or 0-based indexes) to treat as the "
                         "industry classification columns, instead of auto-detecting.")
    return p.parse_args(argv)


def main(argv=None):
    args = parse_args(argv if argv is not None else sys.argv[1:])

    if not args.cas and not args.cas_list:
        raise SystemExit("ERROR: provide CAS numbers via --cas or --cas-list.")
    if not args.harmonised and not args.industry:
        raise SystemExit("ERROR: provide at least one source: --harmonised and/or --industry.")

    cas_inputs = []
    if args.cas:
        cas_inputs.extend(args.cas)
    if args.cas_list:
        cas_inputs.extend(load_cas_list(args.cas_list))
    if not cas_inputs:
        raise SystemExit("ERROR: no CAS numbers found in the provided input.")

    harmonised = None
    if args.harmonised:
        harmonised = HarmonisedIndex(args.harmonised, cas_col_override=args.harmonised_cas_col)
        print_mapping("harmonised", harmonised.mapping)
        print("    rows indexed: {}".format(harmonised.rows_read), file=sys.stderr)

    industry = None
    if args.industry:
        col_overrides = None
        if args.industry_cols:
            col_overrides = [c.strip() for c in args.industry_cols.split(",") if c.strip()]
        industry = IndustryIndex(
            args.industry, cas_col_override=args.industry_cas_col, col_overrides=col_overrides
        )
        print_mapping("industry", industry.mapping)
        print("    rows indexed: {}".format(industry.rows_read), file=sys.stderr)
        if industry.duplicate_cas:
            print("    note: {} CAS number(s) had multiple export rows; kept the first of each."
                  .format(len(industry.duplicate_cas)), file=sys.stderr)

    rows, stats = build_output_rows(cas_inputs, harmonised, industry)
    write_csv(args.output, rows)

    print("\n=== Summary ===", file=sys.stderr)
    print("  CAS numbers processed : {}".format(stats["total"]), file=sys.stderr)
    if stats["no_cas"]:
        print("  no CAS pattern found  : {}".format(stats["no_cas"]), file=sys.stderr)
    if stats["invalid"]:
        print("  failed CAS checksum   : {} (possible typos)".format(stats["invalid"]), file=sys.stderr)
    if harmonised is not None:
        print("  harmonised matches    : {} / {}".format(stats["harmonised"], stats["total"]), file=sys.stderr)
    if industry is not None:
        print("  industry matches      : {} / {}".format(stats["industry"], stats["total"]), file=sys.stderr)
    print("  written               : {}".format(args.output), file=sys.stderr)


if __name__ == "__main__":
    main()
